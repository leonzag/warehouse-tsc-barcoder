from typing import Sequence
from pathlib import Path

import openpyxl as xl

from .typing import (
    ProductDocModel, BoxDocModel,
    LabelType, Data, Dataset, RowNum
)
from barcoder.exceptions import ExcelParsingError

__all__ = ['ExcelParser']

"""
Собирает данные из файла Excel
"""

class ExcelParser:
    """
    Собирает данные из Excel документа в соответствии с типом этикетирования
    """
    def __init__(self, file: Path | str, label_type: LabelType) -> None:
        workbook = xl.load_workbook(file, data_only=True)
        worksheet = workbook.active

        model = {LabelType.BOX: BoxDocModel,
                 LabelType.PRODUCT: ProductDocModel}[label_type]

        if model.start_row >= worksheet.max_row + 1:
            raise ExcelParsingError('The worksheet is empty or does not contain enough lines')

        correct_data, incorrect_data, incorrenct_rows = [], [], []
        for row in range(model.start_row, worksheet.max_row + 1):
            values = (worksheet.cell(row, col).value for col in model.columns)
            data: Data = model.datamaker(**{
                f: v for f, v in zip(model.datamaker._fields, values)
            })
            if all(data):
                correct_data += [data]
            else:
                incorrect_data += [data]
                incorrenct_rows += [row]

        self.__colnames = tuple(model.columns.values())
        self.__correct_data = tuple(correct_data)
        self.__incorrect_data = tuple(incorrect_data)
        self.__incorrect_rows = tuple(incorrenct_rows)

    @property
    def colnames(self) -> Sequence[str]:
        """Имена колонок с данными"""
        return self.__colnames

    @property
    def correct_data(self) -> Dataset:
        """Корректные данные (без пустых ячеек)"""
        return self.__correct_data

    @property
    def incorrect_data(self) -> Dataset:
        """Некоррктные данные (имеются пустные ячейки)"""
        return self.__incorrect_data

    @property
    def incorrenct_rows(self) -> Sequence[RowNum]:
        """Номера рядов с некорректными данными"""
        return self.__incorrect_rows
